"""고급 파일 정리 도구 - 내용 분석·PII 탐지·중복 탐지·감사 로그를 제공한다."""

import re
import json
import shutil
import hashlib
from pathlib import Path
from datetime import datetime
from dataclasses import dataclass, asdict
from typing import Optional, List

# 선택적 import (설치 여부에 따라 텍스트 추출 기능 활성화)
try:
    import fitz  # pymupdf
    _PDF_OK = True
except ImportError:
    _PDF_OK = False

try:
    from docx import Document as DocxDocument
    _DOCX_OK = True
except ImportError:
    _DOCX_OK = False

try:
    from openpyxl import load_workbook
    _XLSX_OK = True
except ImportError:
    _XLSX_OK = False


# 접근 금지 시스템 경로 (CLAUDE.md 보안 규칙)
_BLOCKED_PATHS = [
    "C:\\Windows",
    "C:\\Program Files",
    "C:\\Program Files (x86)",
    "/etc", "/sys", "/proc", "/boot",
]


def _is_safe_path(path: str) -> bool:
    """시스템 보호 경로가 아닌지 확인한다."""
    resolved = str(Path(path).resolve())
    return not any(resolved.startswith(b) for b in _BLOCKED_PATHS)


# ── 결정 데이터 클래스 ──────────────────────────────────────────────────────────

@dataclass
class Decision:
    file_name: str
    file_path: str
    category: str
    confidence: float
    reasons: List[str]
    contains_pii: bool = False       # 프로토타입 버그 수정: contain_pii → contains_pii
    pii_types: Optional[List[str]] = None
    action: str = "move"
    status: str = "approved"
    destination: Optional[str] = None


# ── 파일 정리 에이전트 ──────────────────────────────────────────────────────────

class FileOrganizerAgent:
    """파일 정리 에이전트 — Observe → Reason → Act 루프로 동작한다."""

    # 확장자 → 카테고리 매핑
    EXT_MAP = {
        "이미지":       {".jpg", ".jpeg", ".png", ".gif", ".bmp", ".webp", ".svg", ".ico"},
        "동영상":       {".mp4", ".avi", ".mkv", ".mov", ".wmv", ".flv", ".webm"},
        "음악":         {".mp3", ".wav", ".flac", ".aac", ".ogg", ".m4a"},
        "문서":         {".pdf", ".docx", ".doc", ".txt", ".md", ".hwp", ".rtf"},
        "스프레드시트": {".xlsx", ".xls", ".csv"},
        "프레젠테이션": {".pptx", ".ppt"},
        "압축파일":     {".zip", ".rar", ".7z", ".tar", ".gz", ".bz2"},
        "코드":         {".py", ".js", ".ts", ".html", ".css", ".java", ".c", ".cpp", ".go", ".rs"},
        "실행파일":     {".exe", ".msi", ".bat", ".sh"},
    }

    # 문서 내용 키워드 → 세부 카테고리 (키워드 2개 이상 매칭 시 고신뢰)
    KEYWORD_MAP = {
        "법무_문서":  ["계약", "합의", "서명", "약관", "조항", "법인", "소송"],
        "재무_문서":  ["영수증", "결제", "입금", "세금", "송금", "청구서", "견적"],
        "업무_문서":  ["회의", "프로젝트", "기획", "업무", "보고서", "일정", "미팅"],
        "학습_자료":  ["강의", "학습", "요약", "정리", "문제", "과제", "시험"],
        "개인_문서":  ["일기", "개인", "가족", "메모", "비밀"],
    }

    # PII 탐지 정규식
    PII_PATTERNS = {
        "주민등록번호": re.compile(r"\b\d{6}[-]?[1-4]\d{6}\b"),
        "전화번호":     re.compile(r"\b01[0-9]-?\d{3,4}-?\d{4}\b"),
        "이메일":       re.compile(r"\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}\b"),
        "카드번호":     re.compile(r"\b(?:\d{4}[- ]?){3}\d{4}\b"),
        "계좌번호":     re.compile(r"\b\d{2,4}[- ]?\d{2,6}[- ]?\d{2,6}\b"),
    }

    def __init__(self, target_dir: str):
        if not _is_safe_path(target_dir):
            raise PermissionError(f"접근이 금지된 경로입니다: {target_dir}")

        self.base_path = Path(target_dir)
        if not self.base_path.exists():
            raise FileNotFoundError(f"폴더가 없습니다: {target_dir}")
        if not self.base_path.is_dir():
            raise NotADirectoryError(f"폴더 경로가 아닙니다: {target_dir}")

        # 운영용 폴더 생성
        self.quarantine_dir = self.base_path / "_quarantine"
        self.pending_dir    = self.base_path / "_pending"
        self.log_dir        = self.base_path / "_logs"

        for d in [self.quarantine_dir, self.pending_dir, self.log_dir]:
            d.mkdir(exist_ok=True)

        # 순회 시 제외할 운영 폴더 집합 (프로토타입 버그 수정: quarantine만 제외하던 것을 3개로 확장)
        self._op_dirs = {self.quarantine_dir, self.pending_dir, self.log_dir}

    # ── Observe ─────────────────────────────────────────────────────────────

    def observe(self) -> List[Path]:
        """정리 대상 파일 목록을 반환한다 (운영 폴더 파일 제외)."""
        return [
            f for f in self.base_path.iterdir()
            if f.is_file() and f.parent not in self._op_dirs
        ]

    # ── Reason ──────────────────────────────────────────────────────────────

    def reason(self, file_path: Path) -> Decision:
        """파일을 분석해 이동 결정을 반환한다."""
        ext = file_path.suffix.lower()

        if not ext:
            return Decision(
                file_name=file_path.name,
                file_path=str(file_path),
                category="확장자없음",
                confidence=1.0,
                reasons=["확장자가 없습니다."],
            )

        # 텍스트 추출 가능한 문서 → 내용 기반 분류
        if ext in {".pdf", ".docx", ".xlsx", ".txt", ".md", ".csv"}:
            return self._classify_document(file_path)

        # 나머지 → 확장자 기반 분류
        return Decision(
            file_name=file_path.name,
            file_path=str(file_path),
            category=self._ext_to_category(ext),
            confidence=1.0,
            reasons=[f"확장자 기반 분류: {ext}"],
        )

    def _ext_to_category(self, ext: str) -> str:
        for category, exts in self.EXT_MAP.items():
            if ext in exts:
                return category
        return "기타"

    def _classify_document(self, file_path: Path) -> Decision:
        """문서 내용을 읽어 PII 탐지 → 키워드 분류 순서로 처리한다."""
        text = self._extract_text(file_path)
        pii_types = self._detect_pii(text) if text else []

        # PII 탐지 → quarantine + pending
        if pii_types:
            return Decision(
                file_name=file_path.name,
                file_path=str(file_path),
                category="quarantine",
                confidence=1.0,
                reasons=[f"PII 탐지됨: {', '.join(pii_types)}"],
                contains_pii=True,
                pii_types=pii_types,
                action="quarantine",
                status="pending",
            )

        # 키워드 2개 이상 매칭 → 고신뢰 분류
        if text:
            sample = text[:3000]
            for category, keywords in self.KEYWORD_MAP.items():
                matched = [kw for kw in keywords if kw in sample]
                if len(matched) >= 2:
                    return Decision(
                        file_name=file_path.name,
                        file_path=str(file_path),
                        category=category,
                        confidence=0.8,
                        reasons=[f"키워드 매칭 ({len(matched)}개): {', '.join(matched[:5])}"],
                    )

            # 키워드 1개 → 저신뢰, pending 검토
            for category, keywords in self.KEYWORD_MAP.items():
                matched = [kw for kw in keywords if kw in sample]
                if matched:
                    return Decision(
                        file_name=file_path.name,
                        file_path=str(file_path),
                        category=category,
                        confidence=0.4,
                        reasons=[
                            f"키워드 약한 매칭: {', '.join(matched)}",
                            "신뢰도가 낮아 관리자 검토가 필요합니다.",
                        ],
                        action="pending",
                        status="pending",
                    )

        # 텍스트 없음 → 확장자로 폴백
        return Decision(
            file_name=file_path.name,
            file_path=str(file_path),
            category=self._ext_to_category(file_path.suffix.lower()),
            confidence=0.6,
            reasons=["내용 분류 불가, 확장자로 대체 분류"],
        )

    # ── 텍스트 추출 ──────────────────────────────────────────────────────────

    def _extract_text(self, file_path: Path) -> str:
        """확장자에 맞는 방법으로 문서 텍스트를 추출한다."""
        ext = file_path.suffix.lower()
        try:
            if ext == ".pdf" and _PDF_OK:
                import fitz
                pages = []
                with fitz.open(file_path) as doc:
                    for page in doc:
                        pages.append(page.get_text())
                return "\n".join(pages).strip()

            if ext == ".docx" and _DOCX_OK:
                from docx import Document as D
                return "\n".join(p.text for p in D(file_path).paragraphs).strip()

            if ext == ".xlsx" and _XLSX_OK:
                from openpyxl import load_workbook as lw
                wb = lw(file_path, data_only=True)
                rows = []
                for ws in wb.worksheets:
                    for row in ws.iter_rows(values_only=True):
                        row_text = " ".join(str(c) for c in row if c is not None)
                        if row_text:
                            rows.append(row_text)
                return "\n".join(rows).strip()

            if ext in {".txt", ".md", ".csv"}:
                return file_path.read_text(encoding="utf-8", errors="ignore").strip()

        except Exception:
            pass
        return ""

    # ── PII 탐지 ─────────────────────────────────────────────────────────────

    def _detect_pii(self, text: str) -> List[str]:
        sample = text[:5000]
        return [name for name, pattern in self.PII_PATTERNS.items() if pattern.search(sample)]

    # ── Act ──────────────────────────────────────────────────────────────────

    def act(self, file_path: Path, decision: Decision) -> str:
        """결정에 따라 파일을 이동하고 감사 로그를 기록한다."""
        if decision.action == "pending":
            pending_file = self._save_pending(decision)
            self._write_audit_log(decision)
            return f"PENDING → {pending_file.name}"

        dest_dir = (
            self.quarantine_dir
            if decision.action == "quarantine"
            else self.base_path / decision.category
        )
        dest_dir.mkdir(exist_ok=True)
        dest_file = self._resolve_dest(dest_dir, file_path)

        shutil.move(str(file_path), str(dest_file))
        decision.destination = str(dest_file)
        self._write_audit_log(decision)
        return str(dest_file)

    def _resolve_dest(self, dest_dir: Path, file_path: Path) -> Path:
        """같은 이름 파일이 있으면 _1, _2 … 번호를 붙여 경로를 반환한다."""
        dest = dest_dir / file_path.name
        if not dest.exists():
            return dest
        counter = 1
        while True:
            candidate = dest_dir / f"{file_path.stem}_{counter}{file_path.suffix}"
            if not candidate.exists():
                return candidate
            counter += 1

    # ── 전체 실행 루프 ───────────────────────────────────────────────────────

    def run_once(self) -> dict:
        """폴더 파일을 한 번 순회해 분류·이동한다."""
        files = self.observe()
        if not files:
            return {"결과": "성공", "메시지": "정리할 파일이 없습니다.", "처리 수": 0}

        results = []
        for file_path in files:
            decision = self.reason(file_path)
            dest = self.act(file_path, decision)
            results.append({
                "파일":    file_path.name,
                "카테고리": decision.category,
                "신뢰도":  round(decision.confidence, 3),
                "액션":    decision.action,
                "결과":    dest,
            })

        return {"결과": "성공", "처리 수": len(results), "내역": results}

    # ── 중복 탐지 ────────────────────────────────────────────────────────────

    def find_duplicates(self) -> List[List[dict]]:
        """SHA-256 해시로 내용이 동일한 중복 파일 그룹 목록을 반환한다."""
        hashes: dict = {}
        for file_path in self.observe():
            try:
                hasher = hashlib.sha256()
                with open(file_path, "rb") as f:
                    while chunk := f.read(8192):
                        hasher.update(chunk)
                hashes.setdefault(hasher.hexdigest(), []).append(file_path)
            except (PermissionError, OSError):
                pass

        duplicates = []
        for paths in hashes.values():
            if len(paths) > 1:
                group = []
                for p in paths:
                    stat = p.stat()
                    group.append({
                        "이름":   p.name,
                        "경로":   str(p),
                        "크기":   f"{stat.st_size / 1024:.2f} KB",
                        "수정일": datetime.fromtimestamp(stat.st_mtime).strftime("%Y-%m-%d %H:%M:%S"),
                    })
                duplicates.append(group)
        return duplicates

    def create_duplicate_pending_reports(self) -> List[str]:
        """중복 파일 그룹을 pending JSON으로 저장하고 파일 경로 목록을 반환한다.

        프로토타입 버그 수정: 클래스 외부 독립 함수로 선언되어 있던 것을 클래스 메서드로 이동.
        """
        duplicates = self.find_duplicates()
        created = []
        for group in duplicates:
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            pending_file = self.pending_dir / f"{ts}_duplicate_review.json"
            payload = {
                "type":       "duplicate_review",
                "status":     "pending",
                "created_at": datetime.now().isoformat(timespec="seconds"),
                "files":      group,
                "message":    "중복 파일입니다. 삭제 전 관리자 승인 필요",
            }
            pending_file.write_text(
                json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8"
            )
            created.append(str(pending_file))
        return created

    # ── Pending 승인 ─────────────────────────────────────────────────────────

    def approve_pending_file(self, pending_json_path: str) -> str:
        """pending 파일을 승인해 최종 처리한다."""
        pending_path = Path(pending_json_path)
        if not pending_path.exists():
            return "pending 파일이 없습니다."

        data = json.loads(pending_path.read_text(encoding="utf-8"))

        # 중복 검토 승인
        if data.get("type") == "duplicate_review":
            data["status"] = "approved"
            data["approved_at"] = datetime.now().isoformat(timespec="seconds")
            pending_path.write_text(
                json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8"
            )
            return f"중복 검토 승인 완료: {pending_path.name}"

        # 일반 파일 승인
        file_path = Path(data["file_path"])
        if not file_path.exists():
            return f"원본 파일이 없습니다: {file_path}"

        decision = Decision(
            file_name=data["file_name"],
            file_path=data["file_path"],
            category=(
                "reviewed_sensitive"
                if data["category"] == "quarantine"
                else data["category"]
            ),
            confidence=float(data["confidence"]),
            reasons=data["reasons"] + ["관리자 승인 완료"],
            contains_pii=data.get("contains_pii", False),
            pii_types=data.get("pii_types", []),
            action="quarantine" if data.get("contains_pii") else "move",
            status="approved",
        )
        result = self.act(file_path, decision)
        data.update({
            "status":            "approved",
            "approved_at":       datetime.now().isoformat(timespec="seconds"),
            "final_destination": result,
        })
        pending_path.write_text(
            json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8"
        )
        return f"승인 처리 완료: {result}"

    # ── 감사 로그 ────────────────────────────────────────────────────────────

    def _write_audit_log(self, decision: Decision):
        log_path = self.log_dir / f"audit_{datetime.now().strftime('%Y%m%d')}.jsonl"
        payload = asdict(decision)
        payload["logged_at"] = datetime.now().isoformat(timespec="seconds")
        with log_path.open("a", encoding="utf-8") as f:
            f.write(json.dumps(payload, ensure_ascii=False) + "\n")

    def _save_pending(self, decision: Decision) -> Path:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        pending_file = self.pending_dir / f"{ts}_{Path(decision.file_name).stem}.json"
        pending_file.write_text(
            json.dumps(asdict(decision), ensure_ascii=False, indent=2), encoding="utf-8"
        )
        return pending_file


# ── 에이전트 core.py 연동용 도구 함수 ─────────────────────────────────────────

def organize_folder(path: str) -> dict:
    """폴더 파일을 분류·정리한다. core.py TOOL_HANDLERS에서 호출된다."""
    try:
        return FileOrganizerAgent(path).run_once()
    except (PermissionError, FileNotFoundError, NotADirectoryError) as e:
        return {"결과": "실패", "이유": str(e)}
    except Exception as e:
        return {"결과": "실패", "이유": str(e)}


def find_duplicate_files(path: str) -> dict:
    """중복 파일을 탐지하고 결과를 반환한다. core.py TOOL_HANDLERS에서 호출된다."""
    try:
        duplicates = FileOrganizerAgent(path).find_duplicates()
        if not duplicates:
            return {"결과": "성공", "메시지": "중복 파일이 없습니다.", "그룹 수": 0}
        return {"결과": "성공", "그룹 수": len(duplicates), "중복 그룹": duplicates}
    except Exception as e:
        return {"결과": "실패", "이유": str(e)}


def approve_pending(target_dir: str, pending_json_path: str) -> dict:
    """pending 파일을 승인 처리한다. core.py TOOL_HANDLERS에서 호출된다."""
    try:
        result = FileOrganizerAgent(target_dir).approve_pending_file(pending_json_path)
        return {"결과": "성공", "메시지": result}
    except Exception as e:
        return {"결과": "실패", "이유": str(e)}
